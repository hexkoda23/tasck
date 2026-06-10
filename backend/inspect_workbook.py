
import openpyxl
import sys
from pathlib import Path

wb_path = Path(__file__).parent.parent / "data" / "Copy of Copy of CRM Template.xlsx"

print(f"Checking workbook: {wb_path}")
print(f"Exists: {wb_path.exists()}")

if wb_path.exists():
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"Number of rows: {len(rows)}")
        if len(rows) > 0:
            print(f"Headers (row 1): {rows[0]}")
        if len(rows) > 1:
            print(f"Row 2: {rows[1]}")
        if len(rows) > 2:
            print(f"Row 3: {rows[2]}")
