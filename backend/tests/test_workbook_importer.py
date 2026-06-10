"""Regression tests for the CRM workbook importer."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import v3_workbook_import as wb


def test_workbook_path_exists():
    assert wb.WORKBOOK_PATH.exists(), f"Workbook not found at {wb.WORKBOOK_PATH}"


def test_slug_handles_unicode_and_punctuation():
    assert wb._slug("The Macallan Nigeria!") == "the-macallan-nigeria"
    assert wb._slug("  ") == "x"
    assert wb._slug(None) == "x"


def test_parse_fee_usd_thousands():
    amount, currency = wb._parse_fee("$50k")
    assert amount == 50_000 and currency == "USD"


def test_parse_fee_naira_millions():
    amount, currency = wb._parse_fee("2 million naira")
    assert amount == 2_000_000 and currency == "NGN"


def test_parse_fee_handles_free_and_nil():
    assert wb._parse_fee("free") == (None, None)
    assert wb._parse_fee("Nil") == (None, None)


def test_parse_fee_plain_number():
    assert wb._parse_fee(300000)[0] == 300_000


def test_norm_collapses_whitespace_and_newlines():
    assert wb._norm("  foo\n\n  bar  ") == "foo bar"
    assert wb._norm(None) == ""


def test_framing_stage_canonicalisation():
    assert wb._framing_stage_canon("Framing") == "frame"
    assert wb._framing_stage_canon("Creative Snapshot") == "plan"
    assert wb._framing_stage_canon("Delivery") == "deliver"
    assert wb._framing_stage_canon("Feedback") == "closed"
    assert wb._framing_stage_canon("") == "frame"


def test_likelihood_stage_mapping():
    assert wb._likelihood_to_stage("Project Identified - Move to Framing") == "frame"
    assert wb._likelihood_to_stage("No Immediate Opportunity - Nurture") == "connect"
    assert wb._likelihood_to_stage("") == "connect"


def test_split_list_basic():
    assert wb._split_list("Social Media, Bill Boards, Live Events") == ["Social Media", "Bill Boards", "Live Events"]
    assert wb._split_list("Nil") == []
    assert wb._split_list(None) == []


def test_deterministic_id_stable():
    a = wb._det_id("brand", "coca-cola")
    b = wb._det_id("brand", "coca-cola")
    c = wb._det_id("brand", "pepsi")
    assert a == b
    assert a != c
    assert a.startswith("brand-")


def test_importer_parses_workbook_into_brands():
    importer = wb.WorkbookImporter()
    import openpyxl
    ws_wb = openpyxl.load_workbook(importer.path, data_only=True)
    importer.parse_brands(ws_wb[importer.SHEET_BRANDS])
    assert len(importer.brands) >= 5
    names = {b["company"].lower() for b in importer.brands.values()}
    assert "coca cola" in names
    assert any("all smiles" in n for n in names)
    assert any("open society" in n for n in names)


def test_importer_creates_continuation_contacts():
    importer = wb.WorkbookImporter()
    import openpyxl
    ws_wb = openpyxl.load_workbook(importer.path, data_only=True)
    importer.parse_brands(ws_wb[importer.SHEET_BRANDS])
    # Coca Cola has two contacts: Felix + Zara (continuation row)
    coca = next((b for b in importer.brands.values() if b["company"].lower() == "coca cola"), None)
    assert coca is not None
    contact_names = {c["name"] for c in coca["contacts"]}
    assert "Felix Adejumo" in contact_names
    assert "Zara Uwana" in contact_names


def test_importer_collects_relationship_managers():
    importer = wb.WorkbookImporter()
    import openpyxl
    ws_wb = openpyxl.load_workbook(importer.path, data_only=True)
    importer.parse_brands(ws_wb[importer.SHEET_BRANDS])
    rm_names = {r["name"].lower() for r in importer.rms.values()}
    assert "jennifer" in rm_names
    assert "chioma" in rm_names
    assert "fanii" in rm_names


def test_importer_preserves_seyel_aliases():
    importer = wb.WorkbookImporter()
    importer._upsert_rm("Seyel")
    importer._upsert_rm("seyel")
    seyel = next(r for r in importer.rms.values() if r["normalized_name"] == "seyel")
    assert "Seyel" in seyel["aliases"]
    assert "seyel" in seyel["aliases"]


def test_importer_creators_include_mi_and_kiekie():
    importer = wb.WorkbookImporter()
    import openpyxl
    ws_wb = openpyxl.load_workbook(importer.path, data_only=True)
    importer.parse_creators(ws_wb[importer.SHEET_CREATORS])
    names = {c["name"].lower() for c in importer.creators.values()}
    assert any("mi" in n and "abaga" not in n for n in names) or any("mi abaga" in n for n in names)
    assert any("kiekie" in n for n in names)
