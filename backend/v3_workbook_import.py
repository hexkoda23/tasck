"""TASCK OS — CRM Workbook Importer (clean rewrite).

Reads `backend/data/Copy of Copy of CRM Template.xlsx` and populates:
  v3_brands, v3_contacts, v3_creators, v3_rms, v3_admin_users,
  v3_business_cases, v3_projects, v3_contracts, v3_reports, v3_fees,
  v3_wallet, v3_tasks, v3_insights, v3_templates, v3_meetings

Rules:
- Idempotent: deterministic IDs by slug; re-runs UPDATE never duplicate.
- Continuation-row aware: blank company → attach to last brand as another contact.
- Never wipes destination collections — uses `update_one({id}, {$set: doc}, upsert=True)`.
- Preserves source provenance on every record (workbook, sheet, row, original values).
- Path resolved relative to this file so it works regardless of cwd.
"""
from __future__ import annotations

import hashlib
import logging
import re
import uuid
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger("tasck.v3.workbook_import")

WORKBOOK_FILENAME = "Copy of Copy of CRM Template.xlsx"
WORKBOOK_PATH = Path(__file__).resolve().parent / "data" / WORKBOOK_FILENAME


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _slug(value: Any, fallback: str = "x") -> str:
    s = re.sub(r"[^a-z0-9]+", "-", str(value or fallback).strip().lower()).strip("-")
    return s or fallback


def _det_id(prefix: str, *parts: Any) -> str:
    key = "|".join(str(p or "") for p in parts)
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}-{digest}"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", s)


def _norm_header(value: Any) -> str:
    s = _norm(value).lower()
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def _split_list(value: Any) -> List[str]:
    if value is None:
        return []
    s = _norm(value)
    if not s or s.lower() == "nil":
        return []
    return [p.strip() for p in re.split(r"[;,\u2022\n]+", s) if p.strip()]


def _parse_fee(value: Any) -> Tuple[Optional[float], Optional[str]]:
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None
    s = str(value).strip().lower()
    if not s or s in {"nil", "free", "n/a", "none"}:
        return None, None
    currency = "USD" if ("$" in s or "usd" in s) else ("NGN" if ("₦" in s or "naira" in s or "ngn" in s) else None)
    # Tolerate stray whitespace inside numbers like "$350, 000".
    s_cleaned = re.sub(r"(?<=\d),\s+(?=\d)", ",", s)
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*(k|m|million|thousand)?", s_cleaned)
    if not m:
        return None, currency
    try:
        amount = float(m.group(1).replace(",", ""))
    except ValueError:
        return None, currency
    mult = (m.group(2) or "").lower()
    if mult in {"k", "thousand"}:
        amount *= 1_000
    elif mult in {"m", "million"}:
        amount *= 1_000_000
    return amount, currency


def _likelihood_to_stage(value: Any) -> str:
    s = _norm(value).lower()
    if "no immediate" in s or "nurture" in s:
        return "connect"
    if "project identified" in s or "move to framing" in s:
        return "frame"
    return "connect"


# ---------------------------------------------------------------------------
# Brand resolution from `Framing - Partners` rows.
#
# Each framing row has a `Partner Lead` (person) and rich `Project Context`.
# We never create a brand from these. Instead we resolve the row to a real
# CRM brand using:
#   1. Explicit lead-name → org map (people we know from the workbook).
#   2. Project-context keyword → org slug (e.g. "cocacola" → coca-cola).
# Result is a tuple `(brand_slug, unlinked_org_name)`. If the resolved org
# exists in `CRM - Partners` it becomes `brand_id`; otherwise it falls back
# to `unlinked_brand_name` so the UI can still show a proper company label
# like "NASCO" without polluting `v3_brands`.
# ---------------------------------------------------------------------------

LEAD_TO_ORG: Dict[str, str] = {
    "hamsudeen": "NASCO",
    "zara": "Coca Cola",
    "zara uwana": "Coca Cola",
    "dr chika nwadi": "All Smiles Signature",
    "dr chika": "All Smiles Signature",
    "chika nwadi": "All Smiles Signature",
    "pedro abramovay": "Open Society Foundation",
    "pedro abramovay fanii osf test": "Open Society Foundation",
    "louise ehlers": "OSF",
    "louis ehlers": "Open Society Foundations",
    "eunice baker": "Open Society Foundation",
    "ayisha osori": "Open Sociey Initiatives for West Africa (OSIWA), for OSF",
    "fiona mbambo": "Open Society Foundation",
    "binaifer nowrojie": "OSF",
    "christiana": "CJID",
    "christiana longe": "CJID",
    "christiana longe cjid": "CJID",
    "akintunde babatunde": "CJID",
    "busola ajibola": "CJID",
    "felix adejumo": "Coca Cola",
    "elizabeth anthony": "Pernod Ricard - Chivas",
    "betty": "Pernod Ricard - Chivas",
}

CONTEXT_ORG_KEYWORDS: List[Tuple[str, str]] = [
    ("cocacola", "Coca Cola"),
    ("coca cola", "Coca Cola"),
    ("coca-cola", "Coca Cola"),
    ("all smiles", "All Smiles Signature"),
    ("cjid", "CJID"),
    ("openness index", "CJID"),
    ("osiwa", "Open Sociey Initiatives for West Africa (OSIWA), for OSF"),
    ("nasco", "NASCO"),
    ("cornflakes", "NASCO"),
    ("chivas", "Pernod Ricard - Chivas"),
    ("pernod ricard", "Pernod Ricard - Chivas"),
]

# Lead/context → clean project descriptor.  Order matters (most specific first).
TITLE_KEYWORD_RULES: List[Tuple[str, str]] = [
    ("northern nigeria", "Northern Nigeria Growth Strategy"),
    ("kano", "Northern Nigeria Growth Strategy"),
    ("cocacola", "Northern Nigeria Growth Strategy"),
    ("dentistry", "Social Media Growth Plan"),
    ("all smiles", "Social Media Growth Plan"),
    ("cornflakes", "Cornflakes Influencer Sales Campaign"),
    ("300m boxes", "Cornflakes Influencer Sales Campaign"),
    ("influencer to drive co", "Cornflakes Influencer Sales Campaign"),
    ("openness index", "Openness Index Youth Conversation"),
    ("revive civic engagement", "Youth Civic Engagement Campaign"),
    ("create awareness", "Youth Civic Engagement Campaign"),
    ("creates a first of its kind", "Openness Index Youth Conversation"),
    ("connect  5 super creatives", "Regional Creative Network"),
    ("connect 5 super creatives", "Regional Creative Network"),
    ("commission african creatives", "Pan-African Cultural Festival"),
    ("re-entering nigeria", "Art as Agency Fellowship"),
    ("art as agency", "Art as Agency Fellowship"),
    ("super creatives in the focus", "Cross-Regional Creative Exchange"),
    ("fundable initiative", "Pan-African Cultural Festival"),
    ("chivas", "Relationship Opportunity"),
    ("civic creativity", "Civic Creativity Initiative"),
    ("geopolitical zones", "Civic Creativity Initiative"),
]


def _resolve_org_from_row(lead: str, folder: str, context: str, goal: str) -> Optional[str]:
    """Return canonical organisation name (display label) for a framing row."""
    blob = " ".join([lead, folder, context, goal]).lower()
    lead_k = (lead or "").strip().lower()
    folder_k = (folder or "").strip().lower()
    if lead_k in LEAD_TO_ORG:
        return LEAD_TO_ORG[lead_k]
    if folder_k in LEAD_TO_ORG:
        return LEAD_TO_ORG[folder_k]
    for kw, org in CONTEXT_ORG_KEYWORDS:
        if kw in blob:
            return org
    return None


def _derive_project_descriptor(context: str, goal: str, lead: str, folder: str, fee_raw: str) -> str:
    blob = " ".join([context, goal, lead, folder]).lower()
    for kw, descriptor in TITLE_KEYWORD_RULES:
        if kw in blob:
            return descriptor
    # Fallback: take first 6 words of project_goal, capitalised, or context
    src = (goal or context or "").strip()
    if src:
        words = src.split()[:8]
        return " ".join(w.strip(".,;:") for w in words).strip().capitalize() or "Strategic Opportunity"
    return "Strategic Opportunity"


def _compute_value_display(fee_raw: str, fee_amount: Optional[float], fee_currency: Optional[str],
                           budget_raw: str, budget_amount: Optional[float], budget_currency: Optional[str]):
    """Return (display_amount, display_currency, display_label, engagement_track)."""
    text = (fee_raw or "") + " " + (budget_raw or "")
    low = text.lower()
    # Percentage fees: don't fake a Naira amount
    pct_match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if pct_match:
        return None, None, f"{pct_match.group(1)}% of total project budget", ("grant" if "grant" in low else "paid")
    track = "grant" if any(k in low for k in ("grant", "total grant", "grant policy", "grant for")) else "paid"
    # Prefer fee amount; fall back to budget
    amount = fee_amount if fee_amount else budget_amount
    currency = fee_currency or budget_currency
    if amount and currency == "USD":
        return amount, "USD", None, track
    return amount, currency or ("NGN" if track == "paid" else None), None, track


def _framing_stage_canon(value: Any) -> str:
    s = _norm(value).lower()
    if not s:
        return "frame"
    if "framing" in s:
        return "frame"
    if "creative snapshot" in s or "creative" in s:
        return "plan"
    if "delivery" in s or "deliver" in s:
        return "deliver"
    if "feedback" in s or "report" in s or "closed" in s or "done" in s:
        return "closed"
    return "frame"


def _stable_source_row(values: tuple) -> Dict[str, Any]:
    return {str(i): _norm(v) for i, v in enumerate(values) if v is not None and _norm(v)}


class WorkbookImporter:
    SHEET_BRANDS = "CRM - Partners"
    SHEET_FRAMING = "Framing - Partners"
    SHEET_CREATORS = "CRM - Super Creatives"
    SHEET_CREATIVES_FRAMING = "Super Creatives - Framing"

    def __init__(self, path: Optional[Path] = None):
        self.path = Path(path) if path else WORKBOOK_PATH
        self.batch_id = uuid.uuid4().hex[:10]
        self.warnings: List[str] = []
        self.skipped: List[Dict[str, Any]] = []
        self.brands: Dict[str, Dict[str, Any]] = {}
        self.contacts: Dict[str, Dict[str, Any]] = {}
        self.creators: Dict[str, Dict[str, Any]] = {}
        self.rms: Dict[str, Dict[str, Any]] = {}
        self.rm_alias_to_id: Dict[str, str] = {}
        self.business_cases: Dict[str, Dict[str, Any]] = {}
        self.projects: Dict[str, Dict[str, Any]] = {}
        self.contracts: Dict[str, Dict[str, Any]] = {}
        self.reports: Dict[str, Dict[str, Any]] = {}
        self.fees: Dict[str, Dict[str, Any]] = {}
        self.wallet: Dict[str, Dict[str, Any]] = {}
        self.tasks: Dict[str, Dict[str, Any]] = {}
        self.meetings: Dict[str, Dict[str, Any]] = {}
        self.admin_users: Dict[str, Dict[str, Any]] = {}
        self.templates: Dict[str, Dict[str, Any]] = {}
        self.insights: Dict[str, Dict[str, Any]] = {}
        self.sheet_row_counts: Dict[str, int] = {}

    # --- RM normalisation -------------------------------------------------
    def _upsert_rm(self, raw_name: Any) -> Optional[str]:
        name = _norm(raw_name)
        if not name or name.lower() in {"nil", "none", "n/a"}:
            return None
        alias = name.lower()
        if alias in self.rm_alias_to_id:
            rec = self.rms[self.rm_alias_to_id[alias]]
            if name not in rec["aliases"]:
                rec["aliases"].append(name)
            return rec["id"]
        # canonical = title case of slug (Seyel/Seyelnen both → "seyel" key but preserve aliases)
        canonical = name if any(c.isupper() for c in name[1:]) else name.title()
        canonical = re.sub(r"\s+", " ", canonical).strip()
        rm_id = _det_id("rm", _slug(canonical))
        if rm_id in self.rms:
            rec = self.rms[rm_id]
            if name not in rec["aliases"]:
                rec["aliases"].append(name)
            self.rm_alias_to_id[alias] = rm_id
            return rm_id
        rec = {
            "id": rm_id,
            "name": canonical,
            "normalized_name": _slug(canonical),
            "initials": "".join(w[0] for w in canonical.split()[:2]).upper(),
            "aliases": [name],
            "role": "Relationship Manager",
            "email": None,
            "active": True,
            "source_values": [name],
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
        }
        self.rms[rm_id] = rec
        self.rm_alias_to_id[alias] = rm_id
        return rm_id

    # --- Sheet 1: CRM - Partners -----------------------------------------
    def parse_brands(self, ws) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        headers = [_norm_header(c) for c in rows[1]]

        def col(part: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if part in h:
                    return i
            return None
        c_company = 0
        c_website = col("website")
        c_brand_notes = 2  # first "Notes and Next Actions" (brand-level)
        c_contact = 3
        c_connect = col("connect")
        c_contact_notes = 5  # second "Notes and Next Actions" (contact-level)
        c_role = col("role")
        c_email = col("email")
        c_phone = col("tel") or col("phone")
        c_linkedin = col("linkedin")
        c_rm = col("relationship_manager")
        c_focus = col("key_marketing_focus")
        c_audience = col("primary_target_audience")
        c_channels = col("key_marketing_channels")
        c_kpis = col("marketing_kpis")
        c_desired = col("desired_relationship_status")
        c_likelihood = col("likelihood_to_work_with_tta")

        current_brand_id: Optional[str] = None
        useful_rows = 0
        for ridx, row in enumerate(rows[2:], start=3):
            if not any(c is not None and _norm(c) for c in row):
                continue
            useful_rows += 1
            company = _norm(row[c_company]) if c_company < len(row) else ""
            contact_name = _norm(row[c_contact]) if c_contact < len(row) else ""
            email = _norm(row[c_email]) if c_email is not None and c_email < len(row) else ""

            if not company and not contact_name and not email:
                self.skipped.append({"sheet": self.SHEET_BRANDS, "row": ridx, "reason": "all_key_fields_empty"})
                continue

            if company:
                brand_id = _det_id("brand", _slug(company))
                rm_raw = row[c_rm] if c_rm is not None and c_rm < len(row) else None
                rm_id = self._upsert_rm(rm_raw)
                canonical_rm_name = self.rms[rm_id]["name"] if rm_id else _norm(rm_raw)
                brand = self.brands.get(brand_id) or {
                    "id": brand_id,
                    "company": company,
                    "name": company,
                    "brand_name": company,
                    "website": _norm(row[c_website]) if c_website is not None and c_website < len(row) else "",
                    "primary_contact": contact_name,
                    "role": _norm(row[c_role]) if c_role is not None and c_role < len(row) else "",
                    "email": email,
                    "phone": _norm(row[c_phone]) if c_phone is not None and c_phone < len(row) else "",
                    "linkedin": _norm(row[c_linkedin]) if c_linkedin is not None and c_linkedin < len(row) else "",
                    "status": _norm(row[c_connect]) if c_connect is not None and c_connect < len(row) else "Connecting",
                    "connect_status": _norm(row[c_connect]) if c_connect is not None and c_connect < len(row) else "Connecting",
                    "notes_and_next_actions": _norm(row[c_contact_notes]) if c_contact_notes < len(row) else "",
                    "relationship_manager_name": canonical_rm_name,
                    "source_relationship_manager_name": _norm(rm_raw),
                    "rm_id": rm_id,
                    "key_marketing_focus": _norm(row[c_focus]) if c_focus is not None and c_focus < len(row) else "",
                    "primary_target_audience": _norm(row[c_audience]) if c_audience is not None and c_audience < len(row) else "",
                    "key_marketing_channels": _split_list(row[c_channels]) if c_channels is not None and c_channels < len(row) else [],
                    "marketing_kpis": _norm(row[c_kpis]) if c_kpis is not None and c_kpis < len(row) else "",
                    "desired_relationship_status": _norm(row[c_desired]) if c_desired is not None and c_desired < len(row) else "",
                    "likelihood_to_work_with_tta": _norm(row[c_likelihood]) if c_likelihood is not None and c_likelihood < len(row) else "",
                    "engagement_track_default": "paid",
                    "contacts": [],
                    "source_workbook": WORKBOOK_FILENAME,
                    "source_sheet": self.SHEET_BRANDS,
                    "source_row_number": ridx,
                    "source_original_values": _stable_source_row(row),
                    "created_from_crm_template": True,
                    "import_batch_id": self.batch_id,
                    "imported_at": _now(),
                    "created_at": _now(),
                    "updated_at": _now(),
                }
                self.brands[brand_id] = brand
                current_brand_id = brand_id

            # Always create the contact (primary or continuation)
            if current_brand_id and contact_name:
                rm_raw = row[c_rm] if c_rm is not None and c_rm < len(row) else None
                cid = _det_id("contact", current_brand_id, _slug(contact_name), email)
                is_primary = bool(company) and self.brands[current_brand_id].get("primary_contact") == contact_name
                connect_status = _norm(row[c_connect]) if c_connect is not None and c_connect < len(row) else ""
                contact_notes = _norm(row[c_contact_notes]) if c_contact_notes < len(row) else ""
                contact = {
                    "id": cid,
                    "brand_id": current_brand_id,
                    "name": contact_name,
                    "role": _norm(row[c_role]) if c_role is not None and c_role < len(row) else "",
                    "email": email,
                    "phone": _norm(row[c_phone]) if c_phone is not None and c_phone < len(row) else "",
                    "linkedin": _norm(row[c_linkedin]) if c_linkedin is not None and c_linkedin < len(row) else "",
                    "is_primary": is_primary,
                    "connect_status": connect_status,
                    "notes": contact_notes,
                    "key_marketing_focus": _norm(row[c_focus]) if c_focus is not None and c_focus < len(row) else "",
                    "primary_target_audience": _norm(row[c_audience]) if c_audience is not None and c_audience < len(row) else "",
                    "likelihood_to_work_with_tta": _norm(row[c_likelihood]) if c_likelihood is not None and c_likelihood < len(row) else "",
                    "rm_id": self._upsert_rm(rm_raw),
                    "source_sheet": self.SHEET_BRANDS,
                    "source_row_number": ridx,
                    "source_original_values": _stable_source_row(row),
                    "created_from_crm_template": True,
                    "import_batch_id": self.batch_id,
                    "imported_at": _now(),
                }
                self.contacts[cid] = contact
                existing_contact_ids = {c.get("id") for c in self.brands[current_brand_id]["contacts"]}
                if cid not in existing_contact_ids:
                    self.brands[current_brand_id]["contacts"].append(
                        {"id": cid, "name": contact_name, "role": contact["role"], "email": email, "phone": contact["phone"]}
                    )
                # Derived task
                if contact_notes and len(contact_notes) > 20:
                    self._add_task("brand_note",
                                   f"{self.brands[current_brand_id]['company']}: follow-up",
                                   contact_notes[:500], current_brand_id, contact["rm_id"], ridx)
                # Derived meeting record
                if connect_status and connect_status.lower() not in {"", "nil"}:
                    self._add_meeting(current_brand_id, cid, contact_name, contact["rm_id"],
                                      "qualification" if "connecting" in connect_status.lower() else "follow-up",
                                      connect_status, contact_notes, ridx)
        self.sheet_row_counts[self.SHEET_BRANDS] = useful_rows

    # --- Sheet 2: Framing - Partners --------------------------------------
    def parse_framing(self, ws) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [_norm_header(c) for c in rows[0]]

        def col(part: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if part in h:
                    return i
            return None
        c_folder = col("partner_folder")
        c_lead = col("partner_lead")
        c_stage = col("stage")
        c_notes = col("notes")
        c_context = col("project_context")
        c_goal = col("project_goal")
        c_success = col("success_factors")
        c_framework = col("confirmed_project_framework")
        c_fee = col("tta_engagement_fee")
        c_shortlist = col("creator_shortlist")
        c_budget = col("indicative_budget_range")
        c_scope = col("confirmed_scope")
        c_bc = col("business_case")
        c_agreement = col("agreements_signed")
        c_report = col("project_report")
        c_feedback = col("feedback")
        c_brand_score = col("brand_score")
        c_creative_score = col("creative_score")

        last_folder = ""
        last_lead = ""
        useful = 0
        for ridx, row in enumerate(rows[1:], start=2):
            if not any(c is not None and _norm(c) for c in row):
                continue
            folder = _norm(row[c_folder]) if c_folder is not None and c_folder < len(row) else ""
            lead = _norm(row[c_lead]) if c_lead is not None and c_lead < len(row) else ""
            stage_raw = _norm(row[c_stage]) if c_stage is not None and c_stage < len(row) else ""
            if folder:
                last_folder = folder
            if lead:
                last_lead = lead
            effective_folder = folder or last_folder
            effective_lead = lead or last_lead

            project_context = _norm(row[c_context]) if c_context is not None and c_context < len(row) else ""
            project_goal = _norm(row[c_goal]) if c_goal is not None and c_goal < len(row) else ""
            notes = _norm(row[c_notes]) if c_notes is not None and c_notes < len(row) else ""

            # Skip rows that have neither a project context nor a goal nor any
            # interesting content beyond the stage label — these are visual
            # spacers in the workbook, not real business cases.
            has_substance = bool(project_context or project_goal or notes)
            if not has_substance and not (folder or lead):
                self.skipped.append({"sheet": self.SHEET_FRAMING, "row": ridx, "reason": "blank_row"})
                continue
            if not has_substance and stage_raw and not project_context and not project_goal and not notes:
                # row with only "Framing" stage but no project content
                self.skipped.append({"sheet": self.SHEET_FRAMING, "row": ridx, "reason": "stage_only"})
                continue

            useful += 1

            # CRITICAL RULE: NEVER create a brand from `Framing - Partners`.
            # Resolve the row to a known organisation via the lead-name map and
            # project-context keywords; everything else stays an "unlinked"
            # business case so the UI still has a proper company label
            # without polluting `v3_brands`.
            org_label = _resolve_org_from_row(effective_lead, effective_folder, project_context, project_goal)
            brand_id: Optional[str] = None
            unlinked_brand_name: Optional[str] = None
            if org_label:
                # Try to match the resolved org against existing CRM brands.
                for b in self.brands.values():
                    if b.get("company", "").lower() == org_label.lower():
                        brand_id = b["id"]
                        break
                if not brand_id:
                    unlinked_brand_name = org_label
            else:
                # No resolution at all → store the lead/folder text as a soft label
                unlinked_brand_name = effective_folder or effective_lead or "Unlinked"

            canonical_stage = _framing_stage_canon(stage_raw)
            fee_raw = _norm(row[c_fee]) if c_fee is not None and c_fee < len(row) else ""
            fee_amount, fee_currency = _parse_fee(fee_raw)
            budget_raw = _norm(row[c_budget]) if c_budget is not None and c_budget < len(row) else ""
            budget_amount, budget_currency = _parse_fee(budget_raw)
            shortlist = _split_list(row[c_shortlist]) if c_shortlist is not None and c_shortlist < len(row) else []
            agreement_raw = _norm(row[c_agreement]) if c_agreement is not None and c_agreement < len(row) else ""
            report_raw = _norm(row[c_report]) if c_report is not None and c_report < len(row) else ""
            feedback_raw = _norm(row[c_feedback]) if c_feedback is not None and c_feedback < len(row) else ""

            display_amount, display_currency, value_label, derived_track = _compute_value_display(
                fee_raw, fee_amount, fee_currency, budget_raw, budget_amount, budget_currency,
            )

            # Build a clean title: `Organisation — Project Descriptor`.
            descriptor = _derive_project_descriptor(project_context, project_goal, effective_lead, effective_folder, fee_raw)
            brand_company = self.brands.get(brand_id, {}).get("company") if brand_id else None
            display_company = brand_company or unlinked_brand_name or "Unlinked"
            title = f"{display_company} — {descriptor}" if display_company and descriptor else (descriptor or display_company or f"Business Case #{ridx}")

            bc_id = _det_id("bc", brand_id or f"unlinked-{_slug(unlinked_brand_name or '')}", _slug(descriptor))
            brand = self.brands.get(brand_id, {}) if brand_id else {}
            track = derived_track

            # RM resolution priority: explicit lead canonical name (only if it's
            # a real RM, not a person partner-lead) → brand RM → None.
            rm_id = None
            if effective_lead and effective_lead.lower() not in LEAD_TO_ORG:
                rm_id = self._upsert_rm(effective_lead)
            if not rm_id:
                rm_id = brand.get("rm_id")

            # MERGE: if the same canonical (org, descriptor) already exists,
            # advance to the latest stage and append source-row metadata. The
            # logical opportunity is one business case across stages, not five.
            stage_order = {"connect": 0, "frame": 1, "plan": 2, "deliver": 3, "closed": 4}
            existing = self.business_cases.get(bc_id)
            if existing:
                existing.setdefault("source_rows", [{
                    "sheet": existing.get("source_sheet"),
                    "row": existing.get("source_row_number"),
                }])
                existing["source_rows"].append({"sheet": self.SHEET_FRAMING, "row": ridx})
                if stage_order.get(canonical_stage, -1) > stage_order.get(existing.get("stage", ""), -1):
                    existing["stage"] = canonical_stage
                    existing["stage_label"] = stage_raw or existing.get("stage_label", "")
                    # Mirror the stage advance onto the linked brand-project.
                    proj_id = _det_id("proj", "brand", bc_id)
                    proj = self.projects.get(proj_id)
                    if proj:
                        proj["stage"] = canonical_stage
                        proj["stage_label"] = stage_raw or proj.get("stage_label", "")
                if not existing.get("next_action") and notes:
                    existing["next_action"] = notes
                continue

            self.business_cases[bc_id] = {
                "id": bc_id,
                "brand_id": brand_id,
                "brand_name": display_company,
                "unlinked_brand_name": unlinked_brand_name,
                "partner_lead": effective_lead,
                "partner_folder": effective_folder,
                "title": title,
                "project_descriptor": descriptor,
                "stage": canonical_stage,
                "stage_label": stage_raw,
                "engagement_track": track,
                "engagement_type": track,
                "estimated_value": display_amount or 0,
                "value_amount": display_amount,
                "value_currency": display_currency,
                "value_label": value_label,
                "value_raw": fee_raw or budget_raw,
                "rm_id": rm_id,
                "rm_name": (self.rms[rm_id]["name"] if rm_id and rm_id in self.rms else brand.get("relationship_manager_name") or ""),
                "next_action": notes,
                "health": "on_track",
                "days_in_stage": 0,
                "connect": {
                    "status": brand.get("status") or "Connecting",
                    "intelligence": brand.get("key_marketing_focus") or "",
                    "outreach_angle": "",
                    "suggested_outreach": "",
                },
                "frame": {
                    "project_context": project_context,
                    "project_goal": project_goal,
                    "success_factors": _norm(row[c_success]) if c_success is not None and c_success < len(row) else "",
                    "framework": _norm(row[c_framework]) if c_framework is not None and c_framework < len(row) else "",
                    "alignment_snapshot_status": "approved" if canonical_stage in {"plan", "deliver", "closed"} else None,
                },
                "plan": {
                    "creator_shortlist": shortlist,
                    "confirmed_scope": _norm(row[c_scope]) if c_scope is not None and c_scope < len(row) else "",
                    "business_case_document": _norm(row[c_bc]) if c_bc is not None and c_bc < len(row) else "",
                    "contract_signed_at": _now() if canonical_stage in {"deliver", "closed"} and agreement_raw else None,
                },
                "deliver": {
                    "agreement_status": agreement_raw,
                    "fee_raw": fee_raw, "fee_amount": fee_amount, "fee_currency": fee_currency,
                    "budget_raw": budget_raw, "budget_amount": budget_amount,
                },
                "closure": {
                    "report": report_raw, "feedback": feedback_raw,
                    "brand_score": _norm(row[c_brand_score]) if c_brand_score is not None and c_brand_score < len(row) else "",
                    "creative_score": _norm(row[c_creative_score]) if c_creative_score is not None and c_creative_score < len(row) else "",
                },
                "timeline": [],
                "source_workbook": WORKBOOK_FILENAME,
                "source_sheet": self.SHEET_FRAMING,
                "source_row_number": ridx,
                "source_original_values": _stable_source_row(row),
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id,
                "imported_at": _now(),
                "created_at": _now(),
                "updated_at": _now(),
            }

            # Derived
            self._add_contract(bc_id, brand_id, agreement_raw, canonical_stage, fee_amount, fee_currency, ridx)
            if fee_raw or budget_raw:
                self._add_fee(bc_id, brand_id, fee_raw, fee_amount, fee_currency, budget_raw, budget_amount, budget_currency, ridx)
                self._add_wallet(bc_id, brand_id, fee_raw or budget_raw, fee_amount or budget_amount, fee_currency or budget_currency, ridx)
            if report_raw or feedback_raw or canonical_stage == "closed":
                self._add_report(bc_id, brand_id, title, report_raw, feedback_raw, ridx)
            if canonical_stage in {"frame", "plan"} and not (_norm(row[c_bc]) if c_bc is not None and c_bc < len(row) else ""):
                self._add_task("missing_business_case", f"Draft Business Case: {title}", "Framing/Plan-stage project missing Business Case document.", brand_id, self.business_cases[bc_id]["rm_id"], ridx)
            if canonical_stage in {"deliver", "closed"} and not agreement_raw:
                self._add_task("missing_agreement", f"Send to legal: {title}", "Deliver/closed-stage project without agreement on record.", brand_id, self.business_cases[bc_id]["rm_id"], ridx)
            if canonical_stage == "deliver" and not report_raw:
                self._add_task("missing_report", f"Final report: {title}", "Deliver-stage project without final report.", brand_id, self.business_cases[bc_id]["rm_id"], ridx)

            # Brand-side project record (so the Projects page is populated)
            pid = _det_id("proj", "brand", bc_id)
            self.projects[pid] = {
                "id": pid,
                "title": title,
                "project_descriptor": descriptor,
                "brand_id": brand_id,
                "brand_name": display_company,
                "unlinked_brand_name": unlinked_brand_name,
                "company": display_company,
                "creator_id": None,
                "source_type": "brand_project",
                "business_case_id": bc_id,
                "stage": canonical_stage,
                "stage_label": stage_raw,
                "engagement_track": track,
                "engagement_type": track,
                "relationship_manager_name": (self.rms[rm_id]["name"] if rm_id and rm_id in self.rms else brand.get("relationship_manager_name") or ""),
                "rm_id": rm_id,
                "rm_name": (self.rms[rm_id]["name"] if rm_id and rm_id in self.rms else brand.get("relationship_manager_name") or ""),
                "partner_lead": effective_lead,
                "partner_folder": effective_folder,
                "project_context": project_context,
                "project_goal": project_goal,
                "success_factors": _norm(row[c_success]) if c_success is not None and c_success < len(row) else "",
                "confirmed_framework": _norm(row[c_framework]) if c_framework is not None and c_framework < len(row) else "",
                "confirmed_scope": _norm(row[c_scope]) if c_scope is not None and c_scope < len(row) else "",
                "creator_shortlist": shortlist,
                "budget_raw": budget_raw,
                "budget_amount": budget_amount,
                "budget_currency": budget_currency,
                "fee_raw": fee_raw,
                "fee_amount": fee_amount,
                "fee_currency": fee_currency,
                "value_amount": display_amount,
                "value_currency": display_currency,
                "value_label": value_label,
                "estimated_value": display_amount or 0,
                "next_action": notes,
                "days_in_stage": 0,
                "agreement_status": agreement_raw,
                "report_status": report_raw,
                "feedback_status": feedback_raw,
                "source_workbook": WORKBOOK_FILENAME,
                "source_sheet": self.SHEET_FRAMING,
                "source_row_number": ridx,
                "source_original_values": _stable_source_row(row),
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id,
                "imported_at": _now(),
                "created_at": _now(),
                "updated_at": _now(),
            }
        self.sheet_row_counts[self.SHEET_FRAMING] = useful

    def _match_brand(self, lead: str, folder: str) -> Optional[str]:
        if not lead and not folder:
            return None
        lead_l = lead.lower().strip()
        folder_l = folder.lower().strip()
        for b in self.brands.values():
            if folder_l and folder_l == (b.get("company") or "").lower():
                return b["id"]
            primary = (b.get("primary_contact") or "").lower()
            if lead_l and primary and (lead_l == primary or lead_l in primary or primary.startswith(lead_l)):
                return b["id"]
            for c in b.get("contacts", []):
                cn = (c.get("name") or "").lower()
                if lead_l and cn and (lead_l == cn or lead_l in cn or cn.startswith(lead_l)):
                    return b["id"]
        return None

    # --- Sheet 3: CRM - Super Creatives -----------------------------------
    def parse_creators(self, ws) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        headers = [_norm_header(c) for c in rows[1]]

        def col(part: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if part in h:
                    return i
            return None
        c_name = 0
        c_current = col("current_relationship_status")
        c_desired = col("desired_relationship_status")
        c_website = col("website")
        c_primary = col("primary_contact")
        c_role = col("role")
        c_email = col("email")
        c_phone = col("tel")
        c_linkedin = col("linkedin")
        c_rm = col("relationship_manager")
        c_focus = col("key_marketing_focus")
        c_audience = col("primary_target_audience")
        c_channels = col("key_marketing_channels")
        c_decision = col("decision_making_process")
        c_process = col("current_creative_talent_process")
        c_fee = col("fee_for_engagement")

        useful = 0
        for ridx, row in enumerate(rows[2:], start=3):
            if not any(c is not None and _norm(c) for c in row):
                continue
            name = _norm(row[c_name]) if c_name < len(row) else ""
            if not name:
                self.skipped.append({"sheet": self.SHEET_CREATORS, "row": ridx, "reason": "no_creator_name"})
                continue
            useful += 1
            rm_raw = row[c_rm] if c_rm is not None and c_rm < len(row) else None
            rm_id = self._upsert_rm(rm_raw)
            cid = _det_id("creator", _slug(name))
            fee_raw = _norm(row[c_fee]) if c_fee is not None and c_fee < len(row) else ""
            fee_amount, fee_currency = _parse_fee(fee_raw)
            existing = self.creators.get(cid)
            rec = {
                "id": cid,
                "name": name,
                "creator_name": name,
                "creative_name": name,
                "current_relationship_status": _norm(row[c_current]) if c_current is not None and c_current < len(row) else "",
                "desired_relationship_status": _norm(row[c_desired]) if c_desired is not None and c_desired < len(row) else "",
                "website": _norm(row[c_website]) if c_website is not None and c_website < len(row) else "",
                "primary_contact": _norm(row[c_primary]) if c_primary is not None and c_primary < len(row) else "",
                "role": _norm(row[c_role]) if c_role is not None and c_role < len(row) else "Creator",
                "email": _norm(row[c_email]) if c_email is not None and c_email < len(row) else "",
                "phone": _norm(row[c_phone]) if c_phone is not None and c_phone < len(row) else "",
                "linkedin": _norm(row[c_linkedin]) if c_linkedin is not None and c_linkedin < len(row) else "",
                "relationship_manager_name": _norm(rm_raw),
                "source_relationship_manager_name": _norm(rm_raw),
                "rm_id": rm_id,
                "key_marketing_focus": _norm(row[c_focus]) if c_focus is not None and c_focus < len(row) else "",
                "primary_target_audience": _norm(row[c_audience]) if c_audience is not None and c_audience < len(row) else "",
                "key_marketing_channels": _split_list(row[c_channels]) if c_channels is not None and c_channels < len(row) else [],
                "decision_making_process": _norm(row[c_decision]) if c_decision is not None and c_decision < len(row) else "",
                "current_creative_talent_process": _norm(row[c_process]) if c_process is not None and c_process < len(row) else "",
                "fee_for_engagement_per_month": fee_raw,
                "fee_raw": fee_raw,
                "fee_amount": fee_amount,
                "fee_currency": fee_currency,
                "fee": fee_raw,
                "tier": "Platinum" if (fee_amount and fee_amount >= 30000) else "Gold",
                "genre": _norm(row[c_role]) if c_role is not None and c_role < len(row) else "",
                "source_workbook": WORKBOOK_FILENAME,
                "source_sheet": self.SHEET_CREATORS,
                "source_row_number": ridx,
                "source_original_values": _stable_source_row(row),
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id,
                "imported_at": _now(),
                "created_at": _now(),
                "updated_at": _now(),
                "aliases": [],
            }
            if existing:
                if rm_id and rm_id != existing.get("rm_id"):
                    existing.setdefault("rm_aliases", []).append({"rm_id": rm_id, "row": ridx})
                if not existing.get("fee_amount") and fee_amount:
                    existing["fee_amount"] = fee_amount
                    existing["fee_raw"] = fee_raw
                    existing["fee_currency"] = fee_currency
                continue
            self.creators[cid] = rec
        self.sheet_row_counts[self.SHEET_CREATORS] = useful

    # --- Sheet 4: Super Creatives - Framing -------------------------------
    def parse_creatives_framing(self, ws) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [_norm_header(c) for c in rows[0]]

        def col(part: str) -> Optional[int]:
            for i, h in enumerate(headers):
                if part in h:
                    return i
            return None
        c_folder = 0
        c_project = 1
        c_stage = col("stage")
        c_notes = col("notes")
        c_budget = col("estimated_budget")
        c_lines = col("budget_lines")
        c_agreement = col("agreement")

        last_folder = ""
        useful = 0
        for ridx, row in enumerate(rows[1:], start=2):
            if not any(c is not None and _norm(c) for c in row):
                continue
            folder = _norm(row[c_folder]) if c_folder < len(row) else ""
            project_name = _norm(row[c_project]) if c_project < len(row) else ""
            if folder:
                last_folder = folder
            effective_folder = folder or last_folder
            if not effective_folder and not project_name:
                continue
            useful += 1
            pid = _det_id("proj", _slug(effective_folder), _slug(project_name or f"row{ridx}"))
            stage_raw = _norm(row[c_stage]) if c_stage is not None and c_stage < len(row) else ""
            canonical_stage = _framing_stage_canon(stage_raw)
            budget_raw = _norm(row[c_budget]) if c_budget is not None and c_budget < len(row) else ""
            budget_amount, budget_currency = _parse_fee(budget_raw)
            creator_id = None
            folder_lower = effective_folder.lower()
            for c in self.creators.values():
                if folder_lower and (folder_lower == c["name"].lower() or folder_lower in c["name"].lower()):
                    creator_id = c["id"]
                    break
            creator_value_amount = budget_amount
            creator_value_currency = budget_currency
            self.projects[pid] = {
                "id": pid,
                "title": project_name or effective_folder,
                "project_descriptor": project_name or effective_folder,
                "working_title": not bool(project_name),
                "folder": effective_folder,
                "brand_id": None,
                "brand_name": effective_folder,
                "company": effective_folder,
                "creator_id": creator_id,
                "creator_name": effective_folder,
                "source_type": "creator_project",
                "stage": canonical_stage,
                "stage_label": stage_raw,
                "engagement_track": "direct",
                "engagement_type": "direct",
                "notes": _norm(row[c_notes]) if c_notes is not None and c_notes < len(row) else "",
                "next_action": _norm(row[c_notes]) if c_notes is not None and c_notes < len(row) else "",
                "budget_raw": budget_raw,
                "budget_amount": budget_amount,
                "budget_currency": budget_currency,
                "value_amount": creator_value_amount,
                "value_currency": creator_value_currency,
                "estimated_value": creator_value_amount or 0,
                "days_in_stage": 0,
                "budget_lines": _norm(row[c_lines]) if c_lines is not None and c_lines < len(row) else "",
                "agreement_status": _norm(row[c_agreement]) if c_agreement is not None and c_agreement < len(row) else "",
                "source_workbook": WORKBOOK_FILENAME,
                "source_sheet": self.SHEET_CREATIVES_FRAMING,
                "source_row_number": ridx,
                "source_original_values": _stable_source_row(row),
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id,
                "imported_at": _now(),
                "created_at": _now(),
                "updated_at": _now(),
            }
        self.sheet_row_counts[self.SHEET_CREATIVES_FRAMING] = useful

    # --- Derived records --------------------------------------------------
    def _add_contract(self, bc_id, brand_id, agreement_raw, canonical_stage, fee_amount, fee_currency, ridx):
        cid = _det_id("contract", bc_id)
        if "signed" in agreement_raw.lower() or "agreement" in agreement_raw.lower():
            status = "signed"
        elif "legal" in agreement_raw.lower():
            status = "pending_legal"
        elif canonical_stage == "deliver":
            status = "active"
        elif canonical_stage == "closed":
            status = "completed"
        elif canonical_stage in {"frame", "plan"}:
            status = "draft_needed"
        else:
            status = "not_started"
        self.contracts[cid] = {
            "id": cid,
            "business_case_id": bc_id,
            "brand_id": brand_id,
            "status": status,
            "agreement_description": agreement_raw or "Pending — derived from CRM",
            "value": fee_amount or 0,
            "currency": fee_currency or "NGN",
            "source_sheet": self.SHEET_FRAMING,
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    def _add_fee(self, bc_id, brand_id, fee_raw, fee_amount, fee_currency, budget_raw, budget_amount, budget_currency, ridx):
        fid = _det_id("fee", bc_id)
        is_strategy = "engagement" in (fee_raw or "").lower() or "strategy" in (fee_raw or "").lower()
        self.fees[fid] = {
            "id": fid,
            "business_case_id": bc_id,
            "brand_id": brand_id,
            "type": "strategy_fee" if is_strategy else "project_budget",
            "raw_value": fee_raw or budget_raw,
            "parsed_amount": fee_amount or budget_amount,
            "currency": fee_currency or budget_currency or "NGN",
            "budget_raw": budget_raw,
            "budget_amount": budget_amount,
            "status": "pending_invoice" if fee_amount else "proposed",
            "source_sheet": self.SHEET_FRAMING,
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    def _add_wallet(self, bc_id, brand_id, raw, amount, currency, ridx):
        wid = _det_id("wallet", bc_id)
        self.wallet[wid] = {
            "id": wid,
            "business_case_id": bc_id,
            "brand_id": brand_id,
            "raw_value": raw,
            "amount": amount or 0,
            "currency": currency or "NGN",
            "status": "proposed" if amount else "no_payment_recorded",
            "source_sheet": self.SHEET_FRAMING,
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    def _add_task(self, source, title, description, brand_id, rm_id, ridx):
        tid = _det_id("task", brand_id or "", _slug(title), source)
        self.tasks[tid] = {
            "id": tid,
            "title": title,
            "description": description,
            "source": source,
            "brand_id": brand_id,
            "rm_id": rm_id,
            "status": "open",
            "priority": "P1" if source in {"missing_agreement", "missing_report"} else "P2",
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    def _add_meeting(self, brand_id, contact_id, contact_name, rm_id, meeting_type, status, notes, ridx):
        mid = _det_id("meeting", brand_id, contact_id or "", meeting_type)
        brand = self.brands.get(brand_id, {})
        entity_name = brand.get("company") or brand.get("name") or contact_name or ""
        rm_name = self.rms.get(rm_id, {}).get("name") if rm_id else ""
        # Find a linked business case for this brand to populate business_case_title
        bc_title = ""
        bc_id_link = None
        for bc in self.business_cases.values():
            if bc.get("brand_id") == brand_id:
                bc_title = bc.get("title", "")
                bc_id_link = bc.get("id")
                break
        title_map = {
            "qualification": f"Qualification Call: {entity_name}".strip(": "),
            "connector": f"Business Call (Connector): {entity_name}".strip(": "),
        }
        self.meetings[mid] = {
            "id": mid,
            "title": title_map.get(meeting_type, f"{(meeting_type or 'meeting').title()}: {entity_name}".strip(": ")),
            "meeting_type": meeting_type,
            "type": meeting_type,
            "stage": "before_crm" if meeting_type == "qualification" else "connect",
            "entity_name": entity_name,
            "entity_type": "brand",
            "brand_id": brand_id,
            "business_case_id": bc_id_link,
            "business_case_title": bc_title,
            "contact_id": contact_id,
            "contact_name": contact_name,
            "rm_id": rm_id,
            "rm_name": rm_name,
            "status": status,
            "notes": notes or "",
            "agenda": notes or "",
            "scheduled_for": None,
            "source_sheet": self.SHEET_BRANDS,
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    def _add_report(self, bc_id, brand_id, title, report_raw, feedback_raw, ridx):
        rid = _det_id("report", bc_id)
        self.reports[rid] = {
            "id": rid,
            "business_case_id": bc_id,
            "brand_id": brand_id,
            "title": f"Report: {title}",
            "status": "complete" if report_raw and feedback_raw else "draft",
            "content": report_raw or "(Draft report — derived from CRM. Populate with delivery insights.)",
            "feedback": feedback_raw,
            "source": "crm_derived_test_record" if not report_raw else "crm_template",
            "source_row_number": ridx,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
            "created_at": _now(),
        }

    # --- Admin users + Insights + Templates ------------------------------
    def build_admin_users(self) -> None:
        sid = "admin-super"
        self.admin_users[sid] = {
            "id": sid, "email": "admin@tasck.agency", "name": "Super Admin",
            "role": "super_admin", "active": True,
            "created_from_crm_template": True,
            "import_batch_id": self.batch_id, "imported_at": _now(),
        }
        for rm in self.rms.values():
            uid = f"admin-{rm['normalized_name']}"
            self.admin_users[uid] = {
                "id": uid,
                "email": rm.get("email") or f"{rm['normalized_name']}@tasck.agency",
                "name": rm["name"], "role": "relationship_manager",
                "rm_id": rm["id"], "active": True,
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id, "imported_at": _now(),
            }
        for c in self.contacts.values():
            email = c.get("email", "")
            if email and "@" in email:
                uid = f"user-{_slug(email)}"
                self.admin_users[uid] = {
                    "id": uid, "email": email, "name": c["name"],
                    "role": "brand_contact", "brand_id": c["brand_id"],
                    "active": True,
                    "created_from_crm_template": True,
                    "import_batch_id": self.batch_id, "imported_at": _now(),
                }

    def build_insights(self) -> None:
        brand_status = Counter((b.get("status") or "unknown").strip() for b in self.brands.values())
        creator_status = Counter((c.get("current_relationship_status") or "unknown").strip() for c in self.creators.values())
        rm_counter = Counter(b.get("relationship_manager_name") or "—" for b in self.brands.values())
        bc_stage = Counter(bc["stage"] for bc in self.business_cases.values())

        def _i(kind: str, title: str, payload: Dict[str, Any]) -> None:
            iid = _det_id("insight", kind)
            self.insights[iid] = {
                "id": iid, "kind": kind, "title": title, "payload": payload,
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id, "imported_at": _now(),
                "created_at": _now(),
            }
        _i("brand_status_breakdown", "Brands by connect status", dict(brand_status))
        _i("creator_status_breakdown", "Creators by relationship status", dict(creator_status))
        _i("rm_load", "Relationship-manager workload", dict(rm_counter.most_common(15)))
        _i("business_case_stage", "Business cases by stage", dict(bc_stage))
        _i("contracts_needed", "Projects without signed agreement",
           {"count": sum(1 for c in self.contracts.values() if c["status"] in {"draft_needed", "pending_legal", "not_started"})})
        _i("reports_pending", "Reports still in draft",
           {"count": sum(1 for r in self.reports.values() if r["status"] == "draft")})

    def build_templates(self) -> None:
        templates = [
            ("business-sop", "Business SOP Template", "Connect → Frame → Plan → Deliver → Close"),
            ("alignment-snapshot", "Project Alignment Snapshot Template", "Purpose · Business context · User landscape · Strategic entry point · Strategic direction · Creator approach · Expected outcomes · Commercial context · Why focus matters · Engagement model · Next steps"),
            ("brainstorming", "Brainstorming Template", "Frame the problem · Generate options · Cluster · Score · Decide"),
            ("creative-strategy", "Creative Strategy Snapshot Template", "Executive snapshot · Strategic foundation · Growth plan · Creator strategy · Execution roadmap · Commercial overview · Tracking plan · Risks · Next steps"),
            ("creative-brief", "Creative Brief Template", "Project reference · Context · Role of the creative · Expected scope · Indicative timeline · Working assumptions · Fee indication request · Availability · Confirmation"),
            ("fee-note", "Draft Fee Note Template", "Engagement scope · Coverage · Term duration · Note on separate agency fees"),
            ("brand-feedback", "Brand Feedback Template", "Understanding · Coordination · Representation · Delivery · Overall · Optional comment"),
            ("creative-feedback", "Creative Feedback Template", "Engagement clarity · Representation quality · Coordination · Professionalism · Overall · Optional comment"),
            ("project-report", "Project Report Template", "Campaign · Timeline · Report date · Introduction · Analysis · What worked · What did not · KPIs · Issues · Financial summary · Conclusion"),
            ("sla", "Service Level Agreement Template", "Relationship · Scope · Client obligations · Payment terms · 10% agency fee · Third-party costs · Approvals · Confidentiality · Liability · Termination · Dispute resolution"),
            ("ica", "Independent Creator Agreement Template", "Independence · Agency role · Creator responsibility · Payment routing · 10% agency fee · IP · Confidentiality · Warranties · Force majeure · Termination · Nigerian law"),
            ("service-agreement", "Service Agreement / Vendor Agreement Template", "Scope · Services · Event/project · Payment schedule · Performance · Cancellation · Liability · Force majeure · Signatures"),
        ]
        for slug, name, summary in templates:
            tid = _det_id("template", slug)
            self.templates[tid] = {
                "id": tid, "slug": slug, "name": name, "summary": summary,
                "active": True, "system_template": True,
                "created_from_crm_template": True,
                "import_batch_id": self.batch_id, "imported_at": _now(),
                "created_at": _now(),
            }

    def derive_connect_business_cases(self) -> None:
        """Always emit a Connect-stage discovery record for every CRM brand so
        the Business Case page's Connect filter is never empty. This sits
        alongside any framing-derived business cases the brand may also have."""
        for brand in self.brands.values():
            bc_id = _det_id("bc", brand["id"], "connect")
            if bc_id in self.business_cases:
                continue
            company = brand.get("company") or ""
            self.business_cases[bc_id] = {
                "id": bc_id, "brand_id": brand["id"],
                "brand_name": company,
                "title": f"{company} — Connect",
                "project_descriptor": "Connect",
                "stage": "connect",
                "stage_label": "Connect",
                "engagement_track": brand.get("engagement_track_default", "paid"),
                "engagement_type": brand.get("engagement_track_default", "paid"),
                "estimated_value": 0,
                "value_amount": None,
                "value_currency": None,
                "value_label": None,
                "rm_id": brand.get("rm_id"),
                "rm_name": brand.get("relationship_manager_name", ""),
                "next_action": brand.get("notes_and_next_actions") or "",
                "health": "on_track",
                "days_in_stage": 0,
                "connect": {
                    "status": brand.get("status") or "Connecting",
                    "intelligence": brand.get("key_marketing_focus") or "",
                    "outreach_angle": "", "suggested_outreach": "",
                },
                "frame": {}, "plan": {}, "deliver": {}, "closure": {},
                "timeline": [],
                "source_workbook": WORKBOOK_FILENAME,
                "source_sheet": "(derived from CRM - Partners)",
                "source_row_number": brand.get("source_row_number"),
                "created_from_crm_template": True,
                "derived_from_crm": True,
                "import_batch_id": self.batch_id,
                "imported_at": _now(),
                "created_at": _now(),
                "updated_at": _now(),
            }

    # --- Orchestration ----------------------------------------------------
    @staticmethod
    async def import_all(db) -> Dict[str, Any]:
        """Static entry point (kept for backward compat with existing callers)."""
        importer = WorkbookImporter()
        return await importer._run(db)

    async def _run(self, db) -> Dict[str, Any]:
        if not self.path.exists():
            msg = f"CRM workbook not found at {self.path}. Please upload the file there."
            logger.error(msg)
            return {"success": False, "error": msg, "workbook_path": str(self.path)}

        wb = openpyxl.load_workbook(self.path, data_only=True)
        sheet_names = wb.sheetnames
        logger.info("Loaded workbook %s with sheets: %s", self.path.name, sheet_names)

        if self.SHEET_BRANDS in sheet_names:
            self.parse_brands(wb[self.SHEET_BRANDS])
        if self.SHEET_CREATORS in sheet_names:
            self.parse_creators(wb[self.SHEET_CREATORS])
        if self.SHEET_FRAMING in sheet_names:
            self.parse_framing(wb[self.SHEET_FRAMING])
        if self.SHEET_CREATIVES_FRAMING in sheet_names:
            self.parse_creatives_framing(wb[self.SHEET_CREATIVES_FRAMING])

        self.derive_connect_business_cases()
        self.build_admin_users()
        self.build_insights()
        self.build_templates()

        async def _upsert_many(collection: str, docs: Dict[str, Dict[str, Any]]) -> int:
            for doc in docs.values():
                # Preserve the original `created_at` on re-import so user-created
                # records keep their relative timeline. Only stamp it on first insert.
                doc_for_set = {k: v for k, v in doc.items() if k != "created_at"}
                update_ops: Dict[str, Any] = {"$set": doc_for_set}
                if "created_at" in doc:
                    update_ops["$setOnInsert"] = {"created_at": doc["created_at"]}
                await db[collection].update_one({"id": doc["id"]}, update_ops, upsert=True)
            return len(docs)

        counts = {
            "brands": await _upsert_many("v3_brands", self.brands),
            "contacts": await _upsert_many("v3_contacts", self.contacts),
            "creators": await _upsert_many("v3_creators", self.creators),
            "relationship_managers": await _upsert_many("v3_rms", self.rms),
            "admin_users": await _upsert_many("v3_admin_users", self.admin_users),
            "business_cases": await _upsert_many("v3_business_cases", self.business_cases),
            "projects": await _upsert_many("v3_projects", self.projects),
            "contracts": await _upsert_many("v3_contracts", self.contracts),
            "reports": await _upsert_many("v3_reports", self.reports),
            "fees": await _upsert_many("v3_fees", self.fees),
            "wallet_entries": await _upsert_many("v3_wallet", self.wallet),
            "tasks": await _upsert_many("v3_tasks", self.tasks),
            "insights": await _upsert_many("v3_insights", self.insights),
            "templates": await _upsert_many("v3_templates", self.templates),
            "meetings": await _upsert_many("v3_meetings", self.meetings),
        }
        result = {
            "success": True,
            "workbook_path": str(self.path),
            "sheets_detected": sheet_names,
            "sheet_useful_rows": self.sheet_row_counts,
            "counts": counts,
            "skipped_count": len(self.skipped),
            "skipped_examples": self.skipped[:25],
            "warnings": self.warnings,
            "import_batch_id": self.batch_id,
            "imported_at": _now(),
        }
        logger.info("Workbook import OK: %s", {k: v for k, v in counts.items() if v})
        return result


async def import_crm_workbook(db) -> Dict[str, Any]:
    importer = WorkbookImporter()
    return await importer._run(db)
